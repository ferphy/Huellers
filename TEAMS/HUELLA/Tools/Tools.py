import gc
import shutil
import os.path
import tkinter as tk
from sys import prefix
from time import sleep
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import os
import openpyxl
import datetime
import numpy as np
import re
import sys
import threading

# overview:
# this program takes a txt file with cluster names, filters data from some csv files and outputs the filtered data to new csv files

# the relative input and output data paths are hardcoded here temporarily until we have an interface to select them
OUTPUT_FOLDER = "Output"
CLUSTERS_FILE_PATH = "Data/Clusters.txt" # this file holds the clusters we will output data for
CELLTABLE_FILE_PATH = "" # first step data
QUALITY_BOOST_TRACKING_FILE_PATH = "" # first step data
THOR_FILE_PATH = ""
DB_FOOTPRINT_FILE_PATH = ""
UMTS_3G_FILE_PATH = ""
LTE_4G_FILE_PATH = ""
NR_5G_FILE_PATH = ""
IOM_TEMPLATES_FOLDER = "Data/IOM_Templates" # this folder holds the templates for the IOM data
TEMPLATE_2G = "HUAWEI_GSM_CUSTOMIZED_CELL_.xlsx"
TEMPLATE_3G = "HUAWEI_UMTS_CUSTOMIZED_CELL_.xlsx"
TEMPLATE_4G = "HUAWEI_LTE_CUSTOMIZED_CELL_.xlsx"
TEMPLATE_5G = "HUAWEI_NR_CUSTOMIZED_CELL_.xlsx"
PRB_TEMPLATES_FOLDER = "Data\\PRBs_Templates" # this folder holds the templates for the PRB data
PRB_TEMPLATE = "PRB_THP_.xlsx"
DATA_FOOTPRINT_TEMPLATE_FILE_PATH = "Data\\Data_footprint_template.xlsx"
# Dumb global var until I figure out a non-convoluted way of handling this.
EXPORT_MODE = "default"
USER_CLUSTERS = ""
ERICSSON_LIST = "EMPTY. EXPORT CLUSTERS FIRST BEFORE ATTEMPTING TO VIEW CELLS."
# default date values if not provided by the user 2023-04-17
DATE_START = '2023-04-17 00:00'
DATE_END = datetime.datetime.now().date().strftime('%Y-%m-%d %H:%M')

PRB_TEMPLATE_CHECK = "Data\\PRBs_Templates\\PRB_THP_TEMPLATE_CHECK.xlsx"

class AdapterEricsson:
    _instance = None
    # the purpose of this class is to adapt the input data to the huawei format
    # and sets the input data to the new generated data before generating prb and footprint data
    def __init__(self):
        if AdapterEricsson._instance is not None:
            raise Exception("AdapterEricsson is a singleton class. Use get_instance() method to access the instance.")
        self.input_4g = "Data\\Ericsson\\4G.csv"
        self.input_5g = "Data\\Ericsson\\5G.csv"
        self.input_3g = "Data\\Ericsson\\3G.csv"
        self.output_4g = f'{OUTPUT_FOLDER}/Ericsson/4G_output.csv'
        self.output_5g = f'{OUTPUT_FOLDER}/Ericsson/5G_output.csv'
        self.output_3g = f'{OUTPUT_FOLDER}/Ericsson/3G_output.csv'

    def get_instance():
        if AdapterEricsson._instance is None:
            AdapterEricsson._instance = AdapterEricsson()
        return AdapterEricsson._instance

    def generate_3g_output(self):
        eric_date_key = 'DIA'
        eric_hour_key = 'HORA'
        huawei_date_key = 'Date'
        ericsson_to_huawei_dict = {
            # KPI equivalentes entre Ericsson y Huawei (3G)
            '3G_UTRANCELL': 'Cell Name'
            'VOICE DROP CALL RATE: E3GD003: % RAB Drop Voice': '3G_QF_DCR_Voice(%)',
            'VOICE CALL SETUP SUCCESS RATE FOR FAST DORMANCY: E3GVSS012: % CSSR': '% CSSR CS HW(%)',
            'CALL SETUP SUCCESS RATE PS (R99+HS) FOR FAST DORMANCY: E3GSSPS012: % CSSR PS': '% CSSR PS HW(%)',
            'PS DROP CALL RATE (R99+HS): E3GPSD003: % RAB Drop PS': '3G_QF_DCR_PS(%)',
            'SOFT HANDOVER(EXCL. PREP): E3GSH001 - Voice Sho Success Ratio': '3G_QF_Voice_SHO_Success_Rate(%)',
            'INTERFREQUENCY HARD HANDOVER SUCCESS (EXCL. PREP): E3GHH001: Cs Interfrequency Hard Handover Success Ratio': '3G_QF_PS_HHO_Success_Rate(%)',
            'VOICE IRAT 3G TO 2G: E3GTO2G002: Irat 3g To 2g Voice Handover (Excluding Preparation)': '3G_QF_IRAT_3G_to_2G_Voice_HO (excluding preparation)(%)',
            'THROUGHPUT (KBPS): E3GT002: User Troughput (Kbps)': '3G_QF_DL_Data_Traffic(kB)',
            'THROUGHPUT (KBPS): E3GT003: UL User Throughput (Kbps)': '3G_QF_UL_Data_Traffic(kB)',
            'RSSI 3G: E3GRSSI005: RSSI_NEW': '3G_QF_RSSI_UL(dBm)',

            # Distance bins (TP)
            'TP1 (0.0 - 0.3 Km)': 'VS.TP.UE.0',
            'TP2 (0.3 - 0.7 Km)': 'VS.TP.UE.1',
            'TP3 (0.7 - 1.1 Km)': 'VS.TP.UE.2',
            'TP4 (1.1 - 2.2 Km)': 'VS.TP.UE.3',
            'TP5 (2.2 - 3.7 Km)': 'VS.TP.UE.4',
            'TP6 (3.7 - 6.2 Km)': 'VS.TP.UE.5',
            'TP7 (6.2 - 14.0 Km)': 'VS.TP.UE.6.9',
            'TP8 (>14.0 Km)': 'VS.TP.UE.More55',

            # IRAT interoperability
            'INTEROPERABILITY WITH 4G: E3GI4G001: Cssr Csfb': '3G_QF_Calls ending in 2G(%)'
        }
        try:
            df = pd.read_csv(self.input_3g, delimiter=';')
            # rename the columns to match the huawei format
            df.rename(columns=ericsson_to_huawei_dict, inplace=True)
            # force the eric hour column to the huawei format, HH:MM:SS to HH:MM
            df[eric_hour_key] = pd.to_datetime(df[eric_hour_key], format='%H').dt.strftime('%H:%M')
            df[eric_date_key] = pd.to_datetime(df[eric_date_key], format='%Y%m%d').dt.strftime('%d/%m/%Y')
            # add the huawei date column with the data from the ericsson date and hour columns
            df[huawei_date_key] = pd.to_datetime(df[eric_date_key].astype(str) + ' ' + df[eric_hour_key].astype(str), format='%d/%m/%Y %H:%M')
            # remove the ericsson date and hour columns
            df.drop(columns=[eric_date_key, eric_hour_key], inplace=True)
            # remove duplicate columns to avoid errors, this are in other input data like cell_table or thor_cell_scoring
            other_data_columns = ['SITE']
            for col in other_data_columns:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)
            df.to_csv(self.output_3g, sep=';', index=False)

            #huawei_df = pd.read_csv(self.output_4g, sep=';')
            # append the new columns to the huawei dataframe
            
        except Exception as e:
            print(f"An error occurred while adapting the 3G data from Ericsson: {e}")

    def generate_4g_output(self):
        eric_date_key = 'FECHA'
        eric_hour_key = 'HORA'
        huawei_date_key = 'Date'
        ericsson_to_huawei_dict = {
            'CELLNAME': 'Cell Name',
            'PDCCH USAGE (E4GPD001)': 'PDCCH.Usage.RATE(%)',
            '4G Avg PDCP SDU DL/UL Throughput (Mbps) (E4GTDL001)': '4G_User_DL_Throughput(Mbps)(Mbps)',
            '4G PRB USAGE (E4GPU001)': 'PRB.DL.Usage.RATE(%)', # maybe another column
            'CELL AVAILABILITY (E4GVAIL001)': 'L.ChMeas.PRB.DL.Avail', # maybe another column
            '4G PRB USAGE (E4GPU001) 2': 'L.ChMeas.PRB.DL.Used.Avg' # not found
        }
        try:
            df = pd.read_csv(self.input_4g, delimiter=';')
            df.drop(columns=['SEMANA'], inplace=True)
            # rename the columns to match the huawei format
            df.rename(columns=ericsson_to_huawei_dict, inplace=True)
            # force the eric hour column to the huawei format, HH:MM:SS to HH:MM
            df[eric_hour_key] = pd.to_datetime(df[eric_hour_key], format='%H:%M').dt.strftime('%H:%M')
            # add the huawei date column with the data from the ericsson date and hour columns
            df[huawei_date_key] = pd.to_datetime(df[eric_date_key].astype(str) + ' ' + df[eric_hour_key].astype(str), format='%d/%m/%Y %H:%M')
            # remove the ericsson date and hour columns
            df.drop(columns=[eric_date_key, eric_hour_key], inplace=True)
            # remove duplicate columns to avoid errors, this are in other input data like cell_table or thor_cell_scoring
            other_data_columns = ['SITE']
            for col in other_data_columns:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)
            df.to_csv(self.output_4g, sep=';', index=False)

            #huawei_df = pd.read_csv(self.output_4g, sep=';')
            # append the new columns to the huawei dataframe
            
        except Exception as e:
            print(f"An error occurred while adapting the 4G data from Ericsson: {e}")

    def generate_5g_output(self):
        eric_date_key = 'Dia'
        eric_hour_key = 'HORA'
        huawei_date_key = 'Date'
        ericsson_to_huawei_dict = {
            '5G_GCELDA': 'Cell Name',
            'AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp': 'N.PRB.DL.Avail.Avg', # found
            '5G PRB Use': 'N.PRB.DL.Used.Avg',
        }
        try:
            df = pd.read_csv(self.input_5g, delimiter=';')
            df.drop(columns=['Week'], inplace=True)
            # rename the columns to match the huawei format
            df.rename(columns=ericsson_to_huawei_dict, inplace=True)
            # force the eric hour column to the huawei format, HH:MM:SS to HH:MM
            df[eric_hour_key] = pd.to_datetime(df[eric_hour_key], format='%H:%M:%S').dt.strftime('%H:%M')
            df[eric_date_key] = pd.to_datetime(df[eric_date_key], format='%Y%m%d').dt.strftime('%d/%m/%Y')
            # add the huawei date column with the data from the ericsson date and hour columns
            df[huawei_date_key] = pd.to_datetime(df[eric_date_key].astype(str) + ' ' + df[eric_hour_key].astype(str), format='%d/%m/%Y %H:%M')
            # remove the ericsson date and hour columns
            df.drop(columns=[eric_date_key, eric_hour_key], inplace=True)
            # remove duplicate columns to avoid errors, this are in other input data like cell_table or thor_cell_scoring
            other_data_columns = ['SITE']
            for col in other_data_columns:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)
            df.to_csv(self.output_5g, sep=';', index=False)

        except Exception as e:
            print(f"An error occurred while reading the 5G data from Ericsson: {e}")
            return

    def generate_new_input(self):
        #adapts all the input and sets the input paths to the new generated data
        try:
            if not os.path.exists(f'{OUTPUT_FOLDER}/Ericsson'):
                makedir(f'{OUTPUT_FOLDER}/Ericsson/4G')
            self.generate_4g_output()
            self.generate_5g_output()
            # reset the global input variables to the new generated data
            global LTE_4G_FILE_PATH
            global NR_5G_FILE_PATH
            LTE_4G_FILE_PATH = self.output_4g
            NR_5G_FILE_PATH = self.output_5g

            print(f"New input files generated for ericsson:")
            print(f"4G: {LTE_4G_FILE_PATH}")
            print(f"5G: {NR_5G_FILE_PATH}")
            print(f"3G: {UMTS_3G_FILE_PATH}")
        except Exception as e:
            print(f"An error occurred while generating the new input files: {e}")

class Cluster:
    _instance = None

    def __init__(self):
        if Cluster._instance is not None:
            raise Exception("Cluster is a singleton class. Use get_instance() method to access the instance.")
        self.cluster_names = []

    @staticmethod
    def get_instance():
        if Cluster._instance is None:
            Cluster._instance = Cluster()
        return Cluster._instance

    def get_cluster_names(self):
        try:
            print(f"Validating user input...")
            lines = USER_CLUSTERS.splitlines()
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                self.cluster_names.append(line)
            return True
        except Exception as e:
            print(f"An error occurred while validating user input.: {e}")
            return False

    def reset(self):
        self.cluster_names = []
        print("Cluster instance reset.")

class Ericsson:
    # the output for ericcson is a list of the lites separated by commas
    def __init__(self):
        self.output_file_path = f'{OUTPUT_FOLDER}'
        self.cells_by_cluster = None

    def generate_output(self, cell_table_df, cells_by_cluster_df):
        # generate the output file with the lites separated by commas
        # cells_by_cluster_df has the cluster names and a list of cells named 'CLUSTER_NAME' and 'CELL_LIST'
        key_column = 'VENDOR'
        cell_name_column = 'SITE'
        try:
            eric_df = cell_table_df[cell_table_df['VENDOR'].str.contains('ERICSSON', case=False, na=False)]
            # filter further to leave only the sites, without duplicates
            eric_df = eric_df[[cell_name_column]].drop_duplicates()
            if not os.path.exists(self.output_file_path):
                makedir(self.output_file_path)
            eric_cells_list = eric_df[cell_name_column].tolist()
            global ERICSSON_LIST
            ERICSSON_LIST = ','.join(eric_cells_list)

            #with open(f'{self.output_file_path}/total_ericsson_cells.txt', 'w') as file:
            #    file.write(','.join(eric_cells_list))

            # create a new dataframe by cluster with only the cell names on the eric_df
            # iterate over each cluster list in the cells_by_cluster_df
            # and filter the eric_df by the cell names in the cluster list
            eric_cells_by_cluster = pd.DataFrame(columns=['CLUSTER_NAME', 'CELL_LIST'])
            for index, row in cells_by_cluster_df.iterrows():
                cluster_name = row['CLUSTER_NAME']
                cell_list = row['CELL_LIST'].split(',')
                # filter the eric_df by the cell names in the cell list
                filtered_df = eric_df[eric_df[cell_name_column].isin(cell_list)]
                # join the cell names into a single string separated by commas
                cell_list_str = ','.join(filtered_df[cell_name_column].tolist())
                if cell_list_str:
                    eric_cells_by_cluster = pd.concat([eric_cells_by_cluster, pd.DataFrame({'CLUSTER_NAME': [cluster_name], 'CELL_LIST': [cell_list_str]})], ignore_index=True)
            # save the new dataframe in a txt file in the output folder ericsson_cells_by_cluster.txt
            # each line will have the cluster name and then the cell list separated by commas
            with open(f'{self.output_file_path}/ericsson_cells_by_cluster.txt', 'w') as file:
                for index, row in eric_cells_by_cluster.iterrows():
                    file.write(f"{row['CLUSTER_NAME']}==>{row['CELL_LIST']}\n")

        except Exception as e:
            print(f"An error occurred: {e}")
            return False
    

# CSV Processing ---------------------------------------------------------------

class IOM:
    _instance = None

    def __init__(self):
        if IOM._instance is not None:
            raise Exception("IOM is a singleton class. Use get_instance() method to access the instance.")
        self.output_subfolder = f"{OUTPUT_FOLDER}/IOM"
        # input data
        self.cell_table = None
        self.quality_boost_tracking = None
        # generated
        self.cells_by_cluster = None
        self.cluster_sites = None
        self.cell_table_filtered = None
        self.cell_table_2g = None
        self.cell_table_3g = None
        self.cell_table_4g = None
        self.cell_table_5g = None


    @staticmethod
    def get_instance():
        if IOM._instance is None:
            IOM._instance = IOM()
        return IOM._instance

    def set_quality_boost_tracking(self, quality_boost_tracking_path):
        #sets the quality boost tracking dataframe from the csv in the path
        try:
            self.quality_boost_tracking = pd.read_csv(quality_boost_tracking_path, sep=';')
            print(f"Quality boost tracking data loaded from {quality_boost_tracking_path}.")
        except Exception as e:
            print(f"An error occurred while loading the quality boost tracking data: {e}")

    def set_cell_table(self, cell_table_path):
        # sets the cell table dataframe from the csv in the path
        try:
            self.cell_table = pd.read_csv(cell_table_path, sep=';', low_memory=False) # low_memory=False to avoid dtype warning
            print(f"Cell table data loaded from {cell_table_path}.")
        except Exception as e:
            print(f"An error occurred while loading the cell table data: {e}")

    def set_cells_by_cluster(self):
        # using the cluster sites dataframe, creates a df 
        # the new datafra follow CLUSTER_NAME, CELL_LIST (where cell list has the seed_site joined with the corona_sites list)
        was_successful = True
        try:
            if self.cluster_sites is None:
                self.generate_cluster_sites()
            # create a new dataframe with the cluster names and the cell list
            cluster_names = self.cluster_sites['Cluster'].unique()
            cells_by_cluster = pd.DataFrame(columns=['CLUSTER_NAME', 'CELL_LIST'])
            for cluster_name in cluster_names:
                # get the seed site cell
                seed_site = self.cluster_sites[self.cluster_sites['Cluster'] == cluster_name]['Seed_site'].values[0]
                # get the corona sites cell list
                corona_sites = self.cluster_sites[self.cluster_sites['Cluster'] == cluster_name]['Corona_site'].values[0]
                # join the seed site and corona sites into a single string
                cell_list = f"{seed_site},{corona_sites}"
                # add the cluster name and cell list to the new dataframe
                cells_by_cluster = pd.concat([cells_by_cluster, pd.DataFrame({'CLUSTER_NAME': [cluster_name], 'CELL_LIST': [cell_list]})], ignore_index=True)
            self.cells_by_cluster = cells_by_cluster
            print(f"Cells by cluster dataframe created with {len(self.cells_by_cluster)} clusters.")
            #print the dataframe to the console
            print(self.cells_by_cluster.head())

        except Exception as e:
            print(f"An error occurred while setting the cells by cluster: {e}")
            was_successful = False
        return was_successful


    def generate_cluster_sites(self):
        # generates a csv with CLuster, Seed Site, Corona Site in the output folder
        # using data from the quality boost tracking dataframe
        was_successful = True
        try:
            self.set_quality_boost_tracking(QUALITY_BOOST_TRACKING_FILE_PATH)
            cluster = Cluster.get_instance()
            cluster_names = cluster.cluster_names
            # Filter 'CLUSTER NAME', 'SITES SEMILLA', 'SITES CORONA' from the quality boost tracking dataframe
            filtered_df = self.quality_boost_tracking[
                ['CLUSTER NAME', 'SITES SEMILLA', 'SITES CORONA']
            ]
            #change the column names to 'Cluster', 'Seed_site' and Corona_site'
            filtered_df.columns = ['Cluster', 'Seed_site', 'Corona_site']
            #remove the rows where the cluster name is not in the cluster names list
            filtered_df = filtered_df[filtered_df['Cluster'].isin(cluster_names)]
            self.cluster_sites = filtered_df
            filtered_df.to_csv(f"{OUTPUT_FOLDER}/cluster_sites.csv", sep=';', index=False)

            self.set_cells_by_cluster()
        except Exception as e:
            print(f"An error occurred: {e}")
            was_successful = False
        return was_successful

    def filter_cell_table(self, operator_name, status):
        # filters the cell table dataframe by the operator_name case insensitive, the column is 'OPERATOR NAME'
        # also filters by "STATUS" by the status value
        # then saves the filtered dataframe at self.cell_table_filtered
        was_successful = True
        operator_column = 'OPERATOR NAME'
        cell_name_column = 'SITE'
        status_column = 'STATUS'
        try:
            if self.cell_table is None:
                raise Exception("Cell table is not set.")
            filtered_df = self.cell_table.copy()
            filtered_df = filtered_df[filtered_df[operator_column].str.contains(operator_name, case=False, na=False)]
            filtered_df = filtered_df[filtered_df[status_column] == status]
            # filter to leave relevant data to the input clusters (seed_sites and corona_sites)
            all_seed_sites = self.cluster_sites['Seed_site'].tolist()
            all_corona_sites_list = self.cluster_sites['Corona_site'].tolist()
            all_corona_sites = [site for sublist in all_corona_sites_list for site in sublist.split(',')]
            all_sites = all_seed_sites + all_corona_sites
            filtered_df = filtered_df[filtered_df[cell_name_column].isin(all_sites)]
            huawei_df = filtered_df[filtered_df['VENDOR'].str.contains('HUAWEI', case=False, na=False)]
            eric = Ericsson()
            eric.generate_output(filtered_df, self.cells_by_cluster)

            #self.cell_table_filtered = filtered_df
            self.cell_table_filtered = huawei_df
            return was_successful

        except Exception as e:
            print(f"An error occurred: {e}")
            was_successful = False
        return was_successful

    def generate_cell_table_by_technology(self):
        # creates the 2g, 3g, 4g and 5g cell tables from the filtered cell table
        keyword = 'TECH'
        self.cell_table_2g = self.cell_table_filtered[self.cell_table_filtered[keyword].str.contains('2G', case=False, na=False)]
        self.cell_table_3g = self.cell_table_filtered[self.cell_table_filtered[keyword].str.contains('3G', case=False, na=False)]
        self.cell_table_4g = self.cell_table_filtered[self.cell_table_filtered[keyword].str.contains('4G', case=False, na=False)]
        self.cell_table_5g = self.cell_table_filtered[self.cell_table_filtered[keyword].str.contains('5G', case=False, na=False)]

    def filter_tech_table_by_cluster_sites(self, cluster_name, to_filter_df):
        # return the filtered dataframe by the cluster name
        try:
            result_df = to_filter_df.copy()
            # get the seed site cell
            seed_site = self.cluster_sites[self.cluster_sites['Cluster'] == cluster_name]['Seed_site'].values[0]
            # get the corona sites cell list
            corona_sites = self.cluster_sites[self.cluster_sites['Cluster'] == cluster_name]['Corona_site'].values[0]
            # join the seed site and corona sites into a single string
            cell_list = f"{seed_site},{corona_sites}"
            # filter the dataframe by the cell list
            result_df = result_df[result_df['SITE'].isin(cell_list.split(','))]
            return result_df
        
        except Exception as e:
            print(f"An error occurred while filtering the {cluster_name} table: {e}")
            return None

    def fill_2g_template(self, cluster_name, cell_table_2g_df=None):
        # copies the template file to the output folder and renames it with the cluster name
        # fills the 2g template with the data from the cell table 2g
        # and saves it in the output folder
        keys_to_fill = ['CONTROLLER','CELLNAME','CELLID']
        keys_position = [7, 8, 9] # controller, cellname, cellid respectively in the template
        try:
            if cell_table_2g_df is None:
                raise Exception("Cell table 2G is not set.")
            template_file_path = os.path.join(IOM_TEMPLATES_FOLDER, TEMPLATE_2G)
            # copy the template file to the output folder and add the cluster name to the file name
            if not os.path.exists(self.output_subfolder):
                makedir(self.output_subfolder)
            if not os.path.exists(os.path.join(self.output_subfolder, cluster_name)):
                makedir(os.path.join(self.output_subfolder, cluster_name))
            shutil.copy(template_file_path, os.path.join(self.output_subfolder, cluster_name, TEMPLATE_2G))
            # rename the copied template to have the cluster name at the end (eg HUAWEI_LTE_CUSTOMIZED_CELL_cluster_name.xlsx)
            new_file_name = f"{TEMPLATE_2G.split('.')[0]}{cluster_name}.xlsx"
            new_file_path = os.path.join(self.output_subfolder, cluster_name, new_file_name)
            os.rename(os.path.join(self.output_subfolder, cluster_name, TEMPLATE_2G), new_file_path)
            # open the new file and fill it with the data from the cell table 2g
            fill_template_xlsx(new_file_path, keys_position[0], cell_table_2g_df[keys_to_fill[0]].tolist())
            fill_template_xlsx(new_file_path, keys_position[1], cell_table_2g_df[keys_to_fill[1]].tolist())
            fill_template_xlsx(new_file_path, keys_position[2], cell_table_2g_df[keys_to_fill[2]].tolist())
        except Exception as e:
            print(f"An error occurred: {e}")

    def fill_3g_template(self, cluster_name, cell_table_3g_df=None):
        # copies the template file to the output folder and renames it with the cluster name
        # fills the 3g template with the data from the cell table 3g
        # and saves it in the output folder
        keys_to_fill = ['RNCID','CELLNAME','CELLID']
        keys_position = [7, 8, 9]
        try:
            if cell_table_3g_df is None:
                raise Exception("Cell table 3G is not set.")
            template_file_path = os.path.join(IOM_TEMPLATES_FOLDER, TEMPLATE_3G)
            # copy the template file to the output folder and add the cluster name to the file name
            if not os.path.exists(self.output_subfolder):
                makedir(self.output_subfolder)
            if not os.path.exists(os.path.join(self.output_subfolder, cluster_name)):
                makedir(os.path.join(self.output_subfolder, cluster_name))
            shutil.copy(template_file_path, os.path.join(self.output_subfolder, cluster_name, TEMPLATE_3G))
            # rename the copied template to have the cluster name at the end (eg HUAWEI_LTE_CUSTOMIZED_CELL_cluster_name.xlsx)
            new_file_name = f"{TEMPLATE_3G.split('.')[0]}{cluster_name}.xlsx"
            new_file_path = os.path.join(self.output_subfolder, cluster_name, new_file_name)
            os.rename(os.path.join(self.output_subfolder, cluster_name, TEMPLATE_3G), new_file_path)
            # open the new file and fill it with the data from the cell table 3g
            fill_template_xlsx(new_file_path, keys_position[0], cell_table_3g_df[keys_to_fill[0]].tolist())
            fill_template_xlsx(new_file_path, keys_position[1], cell_table_3g_df[keys_to_fill[1]].tolist())
            fill_template_xlsx(new_file_path, keys_position[2], cell_table_3g_df[keys_to_fill[2]].tolist())
        except Exception as e:
            print(f"An error occurred: {e}")

    def fill_4g_template(self, cluster_name, cell_table_4g_df=None):
        # copies the template file to the output folder and renames it with the cluster name
        # fills the 4g template with the data from the cell table 4g
        # and saves it in the output folder
        keys_to_fill = ['ENODEB_ID', 'CELLNAME', 'CELLID']
        keys_position = [7, 8, 9] # enobed_id, cellname, cellid respectively in the template
        try:
            if cell_table_4g_df is None:
                raise Exception("Cell table 4G is not set.")
            template_file_path = os.path.join(IOM_TEMPLATES_FOLDER, TEMPLATE_4G)
            # copy the template file to the output folder and add the cluster name to the file name
            if not os.path.exists(self.output_subfolder):
                makedir(self.output_subfolder)
            if not os.path.exists(os.path.join(self.output_subfolder, cluster_name)):
                makedir(os.path.join(self.output_subfolder, cluster_name))
            shutil.copy(template_file_path, os.path.join(self.output_subfolder, cluster_name, TEMPLATE_4G))
            # rename the copied template to have the cluster name at the end (eg HUAWEI_LTE_CUSTOMIZED_CELL_cluster_name.xlsx)
            new_file_name = f"{TEMPLATE_4G.split('.')[0]}{cluster_name}.xlsx"
            new_file_path = os.path.join(self.output_subfolder, cluster_name, new_file_name)
            os.rename(os.path.join(self.output_subfolder, cluster_name, TEMPLATE_4G), new_file_path)
            # open the new file and fill it with the data from the cell table 4g
            fill_template_xlsx(new_file_path, keys_position[0], cell_table_4g_df[keys_to_fill[0]].tolist())
            fill_template_xlsx(new_file_path, keys_position[1], cell_table_4g_df[keys_to_fill[1]].tolist())
            fill_template_xlsx(new_file_path, keys_position[2], cell_table_4g_df[keys_to_fill[2]].tolist())
        except Exception as e:
            print(f"An error occurred: {e}")

    def fill_5g_template(self, cluster_name, cell_table_5g_df=None):
        # copies the template file to the output folder and renames it with the cluster name
        # fills the 5g template with the data from the cell table 5g
        # and saves it in the output folder
        keys_to_fill = ['ENODEB_ID', 'CELLNAME', 'CELLID']
        keys_position = [7, 8, 9] # enobed_id, cellname, cellid respectively in the template
        try:
            if cell_table_5g_df is None:
                raise Exception("Cell table 5G is not set.")
            template_file_path = os.path.join(IOM_TEMPLATES_FOLDER, TEMPLATE_5G)
            # copy the template file to the output folder and add the cluster name to the file name
            if not os.path.exists(self.output_subfolder):
                makedir(self.output_subfolder)
            if not os.path.exists(os.path.join(self.output_subfolder, cluster_name)):
                makedir(os.path.join(self.output_subfolder, cluster_name))
            shutil.copy(template_file_path, os.path.join(self.output_subfolder, cluster_name, TEMPLATE_5G))
            # rename the copied template to have the cluster name at the end (eg HUAWEI_LTE_CUSTOMIZED_CELL_cluster_name.xlsx)
            new_file_name = f"{TEMPLATE_5G.split('.')[0]}{cluster_name}.xlsx"
            new_file_path = os.path.join(self.output_subfolder, cluster_name, new_file_name)
            os.rename(os.path.join(self.output_subfolder, cluster_name, TEMPLATE_5G), new_file_path)
            # open the new file and fill it with the data from the cell table 5g
            fill_template_xlsx(new_file_path, keys_position[0], cell_table_5g_df[keys_to_fill[0]].tolist())
            fill_template_xlsx(new_file_path, keys_position[1], cell_table_5g_df[keys_to_fill[1]].tolist())
            fill_template_xlsx(new_file_path, keys_position[2], cell_table_5g_df[keys_to_fill[2]].tolist())
        except Exception as e:
            print(f"An error occurred: {e}")

    def generate_individual_iom_files(self):
        was_successful = True
        # iterate over the cluster names, filter the cell tables by technology for each cluster name
        try:
            for cluster_name in Cluster.get_instance().cluster_names:
                # filter the cell table by the cell names of each cluster
                current_2g_table = self.filter_tech_table_by_cluster_sites(cluster_name, self.cell_table_2g)
                current_3g_table = self.filter_tech_table_by_cluster_sites(cluster_name, self.cell_table_3g)
                current_4g_table = self.filter_tech_table_by_cluster_sites(cluster_name, self.cell_table_4g)
                current_5g_table = self.filter_tech_table_by_cluster_sites(cluster_name, self.cell_table_5g)
                # check if the filtered dataframes are empty
                if (
                    current_4g_table is None or current_4g_table.empty or
                    current_5g_table is None or current_5g_table.empty or
                    current_3g_table is None or current_3g_table.empty or
                    current_2g_table is None or current_2g_table.empty
                ):
                    print(f"{cluster_name} has no huawei cells in the cell table.")
                    was_successful = False
                    continue
                # fill the template with the data from the cell table ng
                self.fill_2g_template(cluster_name, current_2g_table)
                self.fill_3g_template(cluster_name, current_3g_table)
                self.fill_4g_template(cluster_name, current_4g_table)
                self.fill_5g_template(cluster_name, current_5g_table)

        except Exception as e:
            print(f"An error occurred: {e}")
            was_successful = False
        return was_successful

    def generate_tmp_iom_files(self):
        # generates a folder structure for the IOM data for each cluster
        # and populates the folder with the templates from the IOM templates folder
        # using data from the cell table dataframe
        was_successful = True
        try:
            self.set_cell_table(CELLTABLE_FILE_PATH)
            self.generate_cluster_sites()
            self.filter_cell_table("orange", 1)
            self.generate_cell_table_by_technology()
            self.fill_2g_template("ALL", self.cell_table_2g)
            self.fill_3g_template("ALL", self.cell_table_3g)
            self.fill_4g_template("ALL", self.cell_table_4g)
            self.fill_5g_template("ALL", self.cell_table_5g)

            self.generate_individual_iom_files()
        except Exception as e:
            print(f"An error occurred: {e}")
            was_successful = False
        return was_successful

class PRB:
    # singleton class for the PRB data
    _instance = None
    def __init__(self):
        if PRB._instance is not None:
            raise Exception("PRB is a singleton class. Use get_instance() method to access the instance.")
        self.output_subfolder = f"{OUTPUT_FOLDER}/PRB"
        self.df_4g_cleaned = None # temp fix change later
        # input data

    @staticmethod
    def get_instance():
        if PRB._instance is None:
            PRB._instance = PRB()
        return PRB._instance

    def get_input_files(self):
        # loads  QUALITY_BOOST_TRACKIN CELLTABLE THOR_CELL_SCORING DB_FOOTPRINT UMTS_3G LTE_4G NR_5G
        # and saves them in the class variables
        try:
            self.quality_boost_tracking = pd.read_csv(QUALITY_BOOST_TRACKING_FILE_PATH, sep=';')
            self.cell_table = pd.read_csv(CELLTABLE_FILE_PATH, sep=';', low_memory=False) # low_memory=False to avoid dtype warning
            self.thor_cell_scoring = pd.read_csv(THOR_FILE_PATH, sep=';', low_memory=False) # low_memory=False to avoid dtype warning
            self.db_footprint = pd.read_csv(DB_FOOTPRINT_FILE_PATH, sep=';')
            self.umts_3g = pd.read_csv(UMTS_3G_FILE_PATH, sep=';')
            self.lte_4g = pd.read_csv(LTE_4G_FILE_PATH, sep=';')
            self.nr_5g = pd.read_csv(NR_5G_FILE_PATH, sep=';')
            print("PRB input files loaded successfully.")
            return True
        except Exception as e:
            print(f"An error occurred while loading the PRB input files: {e}")
            return False

    def print_input_files(self):
        # prints the input files to the console for debugging purposes
        print("PRB input files:")
        print("QUALITY_BOOST_TRACKING:")
        print(self.quality_boost_tracking.head())
        print("CELLTABLE:")
        print(self.cell_table.head())
        print("THOR_CELL_SCORING:")
        print(self.thor_cell_scoring.head())
        print("DB_FOOTPRINT:")
        print(self.db_footprint.head())
        print("UMTS_3G:")
        print(self.umts_3g.head())
        print("LTE_4G:")
        print(self.lte_4g.head())
        print("NR_5G:")
        print(self.nr_5g.head())

    def get_cell_list(self, cluster_name):
        try:
            iom = IOM.get_instance()
            if iom.cells_by_cluster is None:
                iom.set_cells_by_cluster()
            # get the cell list for this cluster from the IOM previous step
            if cluster_name != "ALL":
                # get the cell list for this cluster from the cells_by_cluster dataframe
                cell_list = IOM.get_instance().cells_by_cluster[IOM.get_instance().cells_by_cluster['CLUSTER_NAME'] == cluster_name]['CELL_LIST'].values[0]
                # creates a list from the cell_list string, ',' as a separator
                cell_list = cell_list.split(',')
                # remove the spaces from the cell names
                cell_list = [cell.strip() for cell in cell_list]
            else:
                # get the cell list for all clusters from the cells_by_cluster dataframe
                cell_list = IOM.get_instance().cells_by_cluster['CELL_LIST'].tolist()
                #join each one of the lists in a single list
                cell_list = [item for sublist in cell_list for item in sublist.split(',')]
            return cell_list
        except Exception as e:
            print(f"An error occurred while getting the cell list for {cluster_name}: {e}")
            return None

    def add_namex_to_dataframe(self, cell_list):
        key_name = "SITE" # for cell_table
        key_namex = "CELLNAMEX" # for cell_table
        key_namex_4g = "Cell Name" # for lte_4g
        key_tech = "TECH" # for cell_table
        try:
            # filter the cell table by the cell names in the cell list
            namex_df = self.cell_table[self.cell_table[key_name].isin(cell_list)]
            # merge the namex_df with the lte_4g dataframe, using the CELLNAMEX and 'Cell Name' columns to match
            namex_df = pd.merge(namex_df, self.lte_4g, left_on=key_namex, right_on=key_namex_4g, how='left')
            # remove the duplicates from the namex_df in the CELLNAMEX column
            namex_df = namex_df.drop_duplicates(subset=[key_namex])
            # remove every column except the CELLNAMEX and the Cell Name and SITE and tech columns
            namex_df = namex_df[[key_name, key_namex, key_namex_4g, key_tech]]
            # remove rows with empty values in the Cell Name column
            namex_df = namex_df[namex_df[key_namex_4g].notna()]
            namex_df.to_csv(os.path.join(OUTPUT_FOLDER, 'namex_df.csv'), sep=';', index=False)
            return namex_df
        except Exception as e:
            print(f"An error occurred while getting the cell namexs: {e}")
            return None

    def add_site_sector_to_dataframe(self, full_df):
        key_sector = 'CELLNAMEX'
        key_site = 'SITE'
        try:
            # Ensure the required columns exist in the dataframe
            if key_sector not in full_df.columns or key_site not in full_df.columns:
                raise KeyError(f"Columns '{key_sector}' or '{key_site}' are missing in the dataframe.")

            # Extract the last integer from CELLNAMEX and combine it with SITE
            full_df['Site-Sector'] = full_df[key_sector].apply(
                lambda x: (
                    f"{full_df.at[full_df.index[full_df[key_sector] == x][0], key_site]}-{match[-1]}"
                    if (match := re.findall(r'\d+', x)) else None
                )
            )
            return full_df
        except Exception as e:
            print(f"An error occurred while adding the Site-Sector column: {e}")
        return None

    def get_site_sector_from_cellname(self, cellname):
        # removes the 4th 10th and 11th characters from the cellname
        # then adds a '-' between the 7th and 8th characters
        # returns the new cellname
        # the cellname is in the format 'ANDX2619W1A' which should be 'AND2619-1'
        try:
            if len(cellname) < 8:
                raise ValueError("Cell name is too short.")
            # remove the 4th, 10th and 11th characters from the cellname
            cellname = cellname[:3] + cellname[4:8] + '-' + cellname[9]
            return cellname
        except Exception as e:
            print(f"An error occurred while getting the site sector from cellname: {e}")
            return None

    def add_site_sector_4g_5g_to_dataframe(self, df):
        # forms the site sector column using the Cell Name column
        # the Cell Name column is in the format 'ANDX2619W1A' which should be 'AND2619-1'
        try:
            # Ensure the required columns exist in the dataframe
            if 'Cell Name' not in df.columns:
                raise KeyError("Column 'Cell Name' is missing in the dataframe.")

            # Extract the site and sector from the Cell Name column
            df['Site-Sector'] = df['Cell Name'].apply(self.get_site_sector_from_cellname)
            df['SITE'] = df['Site-Sector'].apply(lambda x: x.split('-')[0] if '-' in x else None)
            return df
        except Exception as e:
            print(f"An error occurred while adding the Site-Sector column: {e}")

    def classify_4g_bands(self, band):
        # if the BAND column is 'M' or 'Y' then the HB/LB column is 'HB'
        if band in ['M', 'Y']:
            return 'LB'  # Low Band
        else:
            return 'HB'  # High Band

    def add_bands_to_dataframe_4g(self, full_df):
        # using the cellnamex 9th character, adds a new column to the full_df dataframe with the name BAND
        try:
            # create a new column in the full_df dataframe with the name BAND
            full_df['BAND'] = full_df['CELLNAMEX'].str[8:9]
            # create a new column in the full_df dataframe with the name HB/LB (low band or high band)
            full_df['LB/HB'] = full_df['BAND'].apply(self.classify_4g_bands)
            return full_df
        except Exception as e:
            print(f"An error occurred while adding the bands to the dataframe: {e}")
            return None

    def add_pdcch_usage_to_dataframe(self, full_df):
        # gets the highest value in the PDCCH.Usage.RATE(%) column for each cell name in the lte_4g dataframe
        key = 'PDCCH.Usage.RATE(%)'
        try:
            # format the string floats to float values with format_float(value)
            self.lte_4g[key] = self.lte_4g[key].astype(str).str.replace(',', '.')
            #force float conversion of the column
            self.lte_4g[key] = self.lte_4g[key].astype(float)
            # print the type of data in the key column values
            #print(f"Type of data in {key} column: {self.lte_4g[key].dtype}")
            aux_df = self.lte_4g[['Cell Name', key]]
            aux_df = aux_df.groupby('Cell Name', as_index=False)[key].max()
            full_df = pd.merge(full_df, aux_df, left_on='Cell Name', right_on='Cell Name', how='left')
            return full_df
        except Exception as e:
            print(f"An error occurred while getting the max PDCCH usage list: {e}")
            return None

    def add_prb_hc_4g_to_dataframe(self, full_df):
        # gets the highest value in the PRB.DL.Usage.RATE(%) column for each cell name in the lte_4g dataframe
        key = 'PRB.DL.Usage.RATE(%)'
        new_key = 'PRB_HC'
        try:
            # format the string floats to float values with format_float(value)
            self.lte_4g[key] = self.lte_4g[key].astype(str).str.replace(',', '.')
            #force float conversion of the column
            self.lte_4g[key] = self.lte_4g[key].astype(float)
            # print the type of data in the key column values
            #print(f"Type of data in {key} column: {self.lte_4g[key].dtype}")
            # get the max value of the PRB.DL.Usage.RATE(%) column for each cell name in the lte_4g dataframe
            aux_df = self.lte_4g[['Cell Name', key]]
            aux_df = aux_df.groupby('Cell Name', as_index=False)[key].max()
            # rename the column to PRB_HC
            aux_df.rename(columns={key: new_key}, inplace=True)
            # merge the aux_df with the full_df dataframe, using the Cell Name column to match
            full_df = pd.merge(full_df, aux_df, left_on='Cell Name', right_on='Cell Name', how='left')
            return full_df
        except Exception as e:
            print(f"An error occurred while adding the PRB_HC column: {e}")
            return None

    def add_th_hc_to_dataframe(self, full_df):
            # source data from self.lte_4g
            # only uses data between the start and end date
            # if the PRB_HC is >= 70% sets the TH_HC column
            #   calculate the average of all values >= 70 for that cellname
            # if it is not then get the average of all values for that cellname
            key_source = '4G_User_DL_Throughput(Mbps)(Mbps)'
            key_result = 'TH_HC'
            key_mid = 'th_mid'
            key_load = 'th_cargada'
            key_cell = 'Cell Name'
            key_date = 'Date'

            try:
                # Ensure the Date column is in datetime format
                #self.lte_4g[key_date] = pd.to_datetime(self.lte_4g[key_date], format='%Y-%m-%d %H:%M')
                self.lte_4g[key_date] = pd.to_datetime(self.lte_4g[key_date], format='mixed')
                self.lte_4g[key_source] = self.lte_4g[key_source].astype(str).str.replace(',', '.')
                self.lte_4g[key_source] = self.lte_4g[key_source].replace(['/0', ''], np.nan)  # Replace '/0' and empty strings with NaN
                self.lte_4g[key_source] = pd.to_numeric(self.lte_4g[key_source], errors='coerce')  # Convert to numeric, coercing errors to NaN

                # Filter data between the start and end dates
                start_date = pd.to_datetime(DATE_START)
                end_date = pd.to_datetime(DATE_END)
                filtered_data = self.lte_4g[(self.lte_4g[key_date] >= start_date) & (self.lte_4g[key_date] <= end_date)]

                # Initialize columns in the full_df
                full_df[key_mid] = np.nan
                full_df[key_load] = np.nan
                full_df[key_result] = np.nan

                # Process each unique cell name
                for cell in full_df[key_cell].unique():
                    cell_data = filtered_data[filtered_data[key_cell] == cell]

                    # Calculate the average of all values for the cell
                    th_mid_value = np.nanmean(cell_data[key_source])
                    full_df.loc[full_df[key_cell] == cell, key_mid] = th_mid_value

                    # Calculate the average of values >= 70%
                    high_load_data = cell_data[cell_data['PRB.DL.Usage.RATE(%)'].astype(str).str.replace(',', '.').astype(float) >= 70]
                    if not high_load_data.empty:
                        th_cargada_value = np.nanmean(high_load_data[key_source])
                        full_df.loc[full_df[key_cell] == cell, key_load] = th_cargada_value

                    # Set the TH_HC column
                    if not np.isnan(full_df.loc[full_df[key_cell] == cell, key_load].iloc[0]):
                        full_df.loc[full_df[key_cell] == cell, key_result] = full_df.loc[full_df[key_cell] == cell, key_load]
                    else:
                        full_df.loc[full_df[key_cell] == cell, key_result] = th_mid_value

                return full_df

            except Exception as e:
                print(f"An error occurred while adding the TH_HC column: {e}")
                return None

    def add_ok_nok_to_dataframe_4g(self, full_df):
        # checks the values in the TH_HC column and BAND columns
        # then sets the new column OK/NOK to OK or NOK
        general_ok = 6
        YM_ok = 4
        ym = "YM"
        try:
            # create a new column in the full_df dataframe with the name OK/NOK
            full_df['OK/NOK'] = 'NOK'
            # check the values in the TH_HC column and BAND columns
            # if the TH_HC column is >= general_ok and the BAND column is not 'YM' then set the OK/NOK column to OK
            full_df.loc[(full_df['TH_HC'] >= general_ok) & (full_df['BAND'] != ym), 'OK/NOK'] = 'OK'
            # if the TH_HC column is >= YM_ok and the BAND column is 'YM' then set the OK/NOK column to OK
            full_df.loc[(full_df['TH_HC'] >= YM_ok) & (full_df['BAND'] == ym), 'OK/NOK'] = 'OK'
            return full_df
        except Exception as e:
            print(f"An error occurred while adding the OK/NOK column: {e}")
            return None

    def filter_df_by_5g(self, cell_list):
        key_tech = 'TECH'
        try:
            # merge the cell table with the nr_5g dataframe, using the CELLNAMEX and 'Cell Name' columns to match
            # remove the rows where the SITE is not in the cell list

            full_df_5g = self.nr_5g
            # merge with the cell table using the CELLNAME and 'Cell Name' columns to match
            full_df_5g = pd.merge(full_df_5g, self.cell_table, left_on='Cell Name', right_on='CELLNAMEX', how='left')
            full_df_5g = full_df_5g[full_df_5g['SITE'].isin(cell_list)]
            return full_df_5g
        except Exception as e:
            print(f"An error occurred while filtering the 5G dataframe: {e}")
            return None

    def add_degradatio_prb_th_5g(self, full_df_5g):
        # sets the key_result as the % of how much of the available prb is utilized at peak usage
        key_used = 'N.PRB.DL.Used.Avg'
        key_available = 'N.PRB.DL.Avail.Avg'
        key_result = 'DEGRADACION PRB/TH'
        try:
            # format the string with commas to float values
            full_df_5g[key_used] = full_df_5g[key_used].astype(str).str.replace(',', '.')
            full_df_5g[key_available] = full_df_5g[key_available].astype(str).str.replace(',', '.')
            # force float conversion of the columns
            full_df_5g[key_used] = full_df_5g[key_used].astype(float)
            full_df_5g[key_available] = full_df_5g[key_available].astype(float)
            # filter the full_df_5g dataframe to keep only the rows with the max value of key_used for each cell name
            full_df_5g = full_df_5g[full_df_5g[key_used] == full_df_5g.groupby('Cell Name')[key_used].transform('max')].copy()
            # calculate the % of how much of the available prb is utilized at peak usage
            full_df_5g.loc[:, key_result] = (full_df_5g[key_used] / full_df_5g[key_available]) * 100
            #full_df_5g.to_csv(os.path.join(OUTPUT_FOLDER, 'full_df_5g.csv'), sep=';', index=False) # debug
            return full_df_5g
        except Exception as e:
            print(f"An error occurred while adding the DEGRADACION PRB/TH column: {e}")
            return None

    def add_LB_HB_to_dataframe_5g(self, full_df_5g):
        # adds the 'BAND' column to the full_df_5g dataframe
        # using the cellnamex 9th character
        # if the BAND column is 'Q' then the LB/HB column is 'LB'
        # else the LB/HB column is 'HB'
        try:
            # create a new column in the full_df_5g dataframe with the name BAND
            full_df_5g['BAND'] = full_df_5g['Cell Name'].str[8:9]
            # create a new column in the full_df_5g dataframe with the name LB/HB (low band or high band)
            full_df_5g['LB/HB'] = full_df_5g['BAND'].apply(lambda x: 'LB' if x == 'Q' else 'HB')
            return full_df_5g
        except Exception as e:
            print(f"An error occurred while adding the LB/HB column: {e}")
            return None

    def format_input_files_4g_5g(self, lte_4g, nr_5g):
        key_4g_available = 'L.ChMeas.PRB.DL.Avail'
        key_5g_available = 'N.PRB.DL.Avail.Avg'
        key_4g_used = 'L.ChMeas.PRB.DL.Used.Avg'
        key_5g_used = 'N.PRB.DL.Used.Avg'
        key_4g_throughput = '4G_User_DL_Throughput(Mbps)(Mbps)'
        try:
            # format the string with commas to float values
            lte_4g[key_4g_available] = lte_4g[key_4g_available].astype(str).str.replace(',', '.')
            lte_4g[key_4g_used] = lte_4g[key_4g_used].astype(str).str.replace(',', '.')
            lte_4g[key_4g_throughput] = lte_4g[key_4g_throughput].astype(str).str.replace(',', '.')
            nr_5g[key_5g_available] = nr_5g[key_5g_available].astype(str).str.replace(',', '.')
            nr_5g[key_5g_used] = nr_5g[key_5g_used].astype(str).str.replace(',', '.')
            # force float conversion of the columns
            lte_4g[key_4g_available] = lte_4g[key_4g_available].astype(float)
            lte_4g[key_4g_used] = lte_4g[key_4g_used].astype(float)
            lte_4g[key_4g_throughput] = lte_4g[key_4g_throughput].astype(float)
            nr_5g[key_5g_available] = nr_5g[key_5g_available].astype(float)
            nr_5g[key_5g_used] = nr_5g[key_5g_used].astype(float)

            # add the BAND column to the lte_4g dataframe
            lte_4g['BAND'] = lte_4g['Cell Name'].str[8:9]
            # add the BAND column to the nr_5g dataframe
            nr_5g['BAND'] = nr_5g['Cell Name'].str[8:9]
            lte_4g['TECH'] = '4G'
            nr_5g['TECH'] = '5G'
            # print the head of the dataframes BAND and Cell Name columns
            #print(f"lte_4g BAND column:\n{lte_4g[['Cell Name', 'BAND']].head()}")
            #print(f"nr_5g BAND column:\n{nr_5g[['Cell Name', 'BAND']].head()}")

            return lte_4g, nr_5g

        except Exception as e:
            print(f"An error occurred while formatting the input files: {e}")
            return None

    def band_5g_to_4g(self, band):
        matches = {'Q':'Y','W':'T','X':'L'}
        if band in matches:
            return matches[band]
        else:
            return ''

    def filter_df_4g_bands(self, full_df_4g_5g):
        # removes the rows where the TECH column is '4G' and the BAND does not have a 5G equivalent
        try:
            # create a new column in the full_df_4g_5g dataframe with the name BAND_5G
            full_df_4g_5g['BAND_5G'] = full_df_4g_5g['BAND'].apply(self.band_4g_to_5g)
            # remove the rows where the TECH column is '4G' and the BAND does not have a 5G equivalent
            full_df_4g_5g = full_df_4g_5g[~((full_df_4g_5g['TECH'] == '4G') & (full_df_4g_5g['BAND'].isin(['M', 'Y'])))]
            return full_df_4g_5g

        except:
            print(f"An error occurred while filtering the 4G bands: {np.e}")
            return None

    def filter_df_by_4g_5g(self, cell_list):
        # formats the input files, creates the empty dataframe full_df_4g_5g
        # the adds the site sector values with the corresponding BAND and tech for each
        try:
            # format the input files
            copy_df_4g = self.lte_4g.copy()
            copy_df_5g = self.nr_5g.copy()

            copy_df_4g, copy_df_5g = self.format_input_files_4g_5g(copy_df_4g, copy_df_5g)

            copy_df_4g = self.add_site_sector_4g_5g_to_dataframe(copy_df_4g)
            copy_df_5g = self.add_site_sector_4g_5g_to_dataframe(copy_df_5g)

            # remove the rows where the SITE is not in the cell list
            copy_df_4g = copy_df_4g[copy_df_4g['SITE'].isin(cell_list)]
            copy_df_5g = copy_df_5g[copy_df_5g['SITE'].isin(cell_list)]
            self.df_4g_cleaned = copy_df_4g.copy() # temp fix, change later
            print(f"copy_df_4g:\n{copy_df_4g[['Cell Name', 'Site-Sector', 'BAND']].head()}")
            print(f"copy_df_5g:\n{copy_df_5g[['Cell Name', 'Site-Sector', 'BAND']].head()}")

            # new empty dataframe to store the results, add the unique values of the Site-Sector column
            full_df_4g_5g = pd.DataFrame(columns=['Site-Sector', 'BANDS', 'PRB_SUM', '4G_TH_HC', 'OK/NOK', 'TECH'])
            print(f"full_df_4g_5g:\n{full_df_4g_5g.head()}")

            full_df_4g_5g.to_csv(os.path.join(OUTPUT_FOLDER, f'{cell_list[0]}_full_df_4g_5g.csv'), sep=';', index=False) # debug
            full_df_4g_5g = self.process_4g_5g_dataframes(full_df_4g_5g, copy_df_4g, copy_df_5g)
            return full_df_4g_5g
        except Exception as e:
            print(f"An error occurred while filtering the 4G and 5G dataframes: {e}")
            return None

    def process_band_4g_5g(self, full_df_4g_5g, site_sector, bands_4g, current_band_5g, copy_df_4g, copy_df_5g):
        key_npr_available = 'N.PRB.DL.Avail.Avg'
        key_npr_used = 'N.PRB.DL.Used.Avg'
        key_4g_available = 'L.ChMeas.PRB.DL.Avail'
        key_4g_used = 'L.ChMeas.PRB.DL.Used.Avg'
        key_throughput = '4G_User_DL_Throughput(Mbps)(Mbps)'
        key_prb_sum = 'PRB_SUM'
        try:
            # take the values of this site sector and band from the 5g dataframe
            npr_available_5g_list = copy_df_5g[(copy_df_5g['Site-Sector'] == site_sector) & (copy_df_5g['BAND'] == current_band_5g)][key_npr_available].tolist()
            npr_used_5g_list = copy_df_5g[(copy_df_5g['Site-Sector'] == site_sector) & (copy_df_5g['BAND'] == current_band_5g)][key_npr_used].tolist()
            aux_4g_band = self.band_5g_to_4g(current_band_5g)
            if aux_4g_band in bands_4g:
                band_values = current_band_5g + aux_4g_band
                full_df_4g_5g = pd.concat([full_df_4g_5g, pd.DataFrame([{'Site-Sector': site_sector, 'BANDS': band_values, 'PRB_SUM': 0, '4G_TH_HC': 0, 'OK/NOK': 'NOK', 'TECH': '4G'}])], ignore_index=True)
                npr_available_4g_list = copy_df_4g[(copy_df_4g['Site-Sector'] == site_sector) & (copy_df_4g['BAND'] == aux_4g_band)][key_4g_available].tolist()
                npr_used_4g_list = copy_df_4g[(copy_df_4g['Site-Sector'] == site_sector) & (copy_df_4g['BAND'] == aux_4g_band)][key_4g_used].tolist()
                throughput_4g_list = copy_df_4g[(copy_df_4g['Site-Sector'] == site_sector) & (copy_df_4g['BAND'] == aux_4g_band)][key_throughput].tolist()
                if len(np.unique(throughput_4g_list)) > 1:
                    prb_sum = (np.nanmean(npr_used_4g_list) + np.nanmean(npr_used_5g_list)) / (np.nanmean(npr_available_4g_list) + np.nanmean(npr_available_5g_list))
                    full_df_4g_5g[key_prb_sum] = full_df_4g_5g[key_prb_sum].astype(float)
                    full_df_4g_5g.loc[(full_df_4g_5g['Site-Sector'] == site_sector) & (full_df_4g_5g['BANDS'] == band_values), key_prb_sum] = float(prb_sum)
                    if prb_sum > 70:
                        # reduce the throughput list to the values that are in the position of a used greater than 70
                        throughput_4g_list = [throughput_4g_list[i] for i in range(len(throughput_4g_list)) if npr_used_4g_list[i] > 70]
                    full_df_4g_5g.loc[(full_df_4g_5g['Site-Sector'] == site_sector) & (full_df_4g_5g['BANDS'] == band_values), '4G_TH_HC'] = np.nanmean(throughput_4g_list)
                    limit = 6
                    if current_band_5g == 'Q':
                        limit = 4
                    if (full_df_4g_5g.loc[(full_df_4g_5g['Site-Sector'] == site_sector) & (full_df_4g_5g['BANDS'] == band_values), '4G_TH_HC'] >= limit).any():
                        full_df_4g_5g.loc[(full_df_4g_5g['Site-Sector'] == site_sector) & (full_df_4g_5g['BANDS'] == band_values), 'OK/NOK'] = 'OK'
            else:
                prb_sum = (np.nanmean(npr_used_5g_list)) / (np.nanmean(npr_available_5g_list))
                full_df_4g_5g = pd.concat([full_df_4g_5g, pd.DataFrame([{'Site-Sector': site_sector, 'BANDS': current_band_5g, 'PRB_SUM': prb_sum, '4G_TH_HC': 0, 'OK/NOK': 'NOK', 'TECH': '5G'}])], ignore_index=True)
                # if the prb sum is smaller than 70 then set the OK/NOK column to ok
                if prb_sum <= 70:
                    full_df_4g_5g.loc[(full_df_4g_5g['Site-Sector'] == site_sector) & (full_df_4g_5g['BANDS'] == current_band_5g), 'OK/NOK'] = 'OK'

            return full_df_4g_5g
        except Exception as e:
            print(f"An error occurred while processing the band {current_band_5g} for site sector {site_sector}: {e}")
            return None

    def process_4g_5g_dataframes(self, full_df_4g_5g, copy_df_4g, copy_df_5g):
        try:
            site_sectors = copy_df_5g['Site-Sector'].unique()
            for site_sector in site_sectors:
                bands_4g = np.unique(copy_df_4g[copy_df_4g['Site-Sector'] == site_sector]['BAND'])
                # take the 5g bands except the 'P' band
                bands_5g = np.unique(copy_df_5g[copy_df_5g['Site-Sector'] == site_sector]['BAND'])
                bands_5g = [band for band in bands_5g if band != 'P']
                for band in bands_5g:
                    full_df_4g_5g = self.process_band_4g_5g(full_df_4g_5g, site_sector, bands_4g, band, copy_df_4g, copy_df_5g)
            return full_df_4g_5g
        except Exception as e:
            print(f"An error occurred while processing the 4G and 5G dataframes: {e}")
            return None

    def add_node_to_dataframe(self, df_node):
        # adds a NODE column to the dataframe
        # with the value of the Cell Name without the last 3 characters
        try:
            # Ensure the required columns exist in the dataframe
            if 'Cell Name' not in df_node.columns:
                raise KeyError("Column 'Cell Name' is missing in the dataframe.")
            df_node['NODE'] = df_node['Cell Name'].str[:-3]
            return df_node
        except Exception as e:
            print(f"An error occurred while adding the NODE column: {e}")
            return None

    def filter_balance_df(self, cell_list):
        # uses the cell table, lte_4g
        # fills the bands columns with the mean prb usage for that site sector and band
        balance_columns = ['ENODEB','SITE-SECTOR','Y','M','N','T','L']
        balance_df = pd.DataFrame(columns=balance_columns)
        key_prb = 'PRB.DL.Usage.RATE(%)'
        key_enodeb = 'NODE'
        try:
            copy_df_4g = self.df_4g_cleaned.copy()
            copy_df_4g = self.add_node_to_dataframe(copy_df_4g)
            print(f"copy_df_4g:\n{copy_df_4g[['NODE','Cell Name', 'Site-Sector', 'BAND']].head()}")
            site_sector_list = copy_df_4g['Site-Sector'].unique()
            for site_sector in site_sector_list:
                site_sector_node = copy_df_4g[copy_df_4g['Site-Sector'] == site_sector][key_enodeb].tolist()
                balance_df = pd.concat([balance_df, pd.DataFrame([{'ENODEB': site_sector_node[0], 'Site-Sector': site_sector, 'Y': 0, 'M': 0, 'N': 0, 'T': 0, 'L': 0}])], ignore_index=True)
                band_list = np.unique(copy_df_4g[copy_df_4g['Site-Sector'] == site_sector]['BAND'])
                for band in band_list:
                    prb_usage_list = copy_df_4g[(copy_df_4g['Site-Sector'] == site_sector) & (copy_df_4g['BAND'] == band)][key_prb].tolist()
                    prb_usage_mean = np.nanmean(prb_usage_list)
                    # set the value of the band in the balance_df dataframe
                    balance_df.loc[(balance_df['ENODEB'] == site_sector_node[0]) & (balance_df['Site-Sector'] == site_sector), band] = prb_usage_mean
            return balance_df
        except Exception as e:
            print(f"An error occurred while filtering the balance dataframe: {e}")
            return None

    def fill_cluster_dfs(self, template_file_path, cluster_name):
        try:
            print(f"Filling PRB template for {cluster_name}...")
            cell_list = self.get_cell_list(cluster_name)
            full_df_4g = self.add_namex_to_dataframe(cell_list)
            full_df_4g = self.add_pdcch_usage_to_dataframe(full_df_4g)
            full_df_4g = self.add_bands_to_dataframe_4g(full_df_4g)
            full_df_4g = self.add_prb_hc_4g_to_dataframe(full_df_4g)
            full_df_4g = self.add_th_hc_to_dataframe(full_df_4g)
            full_df_4g = self.add_ok_nok_to_dataframe_4g(full_df_4g)

            fill_template_xlsx_row(template_file_path, 1, full_df_4g['CELLNAMEX'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 2, full_df_4g['LB/HB'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 3, full_df_4g['PDCCH.Usage.RATE(%)'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 4, full_df_4g['PRB_HC'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 5, full_df_4g['TH_HC'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 6, full_df_4g['OK/NOK'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 8, full_df_4g['th_mid'].tolist(), 6)
            fill_template_xlsx_row(template_file_path, 9, full_df_4g['th_cargada'].tolist(), 6)

            full_df_5g = self.filter_df_by_5g(cell_list)
            full_df_5g = self.add_degradatio_prb_th_5g(full_df_5g)
            full_df_5g = self.add_LB_HB_to_dataframe_5g(full_df_5g)

            fill_template_xlsx_row(template_file_path, 17, full_df_5g['Cell Name'].tolist(), 4)
            fill_template_xlsx_row(template_file_path, 18, full_df_5g['LB/HB'].tolist(), 4)
            fill_template_xlsx_row(template_file_path, 19, full_df_5g['DEGRADACION PRB/TH'].tolist(), 4)

            full_df_4g_5g = self.filter_df_by_4g_5g(cell_list)

            fill_template_xlsx_row(template_file_path, 11, full_df_4g_5g['Site-Sector'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 12, full_df_4g_5g['BANDS'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 13, full_df_4g_5g['PRB_SUM'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 14, full_df_4g_5g['4G_TH_HC'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 15, full_df_4g_5g['OK/NOK'].tolist(), 5)

            full_df_balance = self.filter_balance_df(cell_list)

            fill_template_xlsx_row(template_file_path, 26, full_df_balance['ENODEB'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 27, full_df_balance['Site-Sector'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 25, full_df_balance['L'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 24, full_df_balance['T'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 23, full_df_balance['N'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 22, full_df_balance['M'].tolist(), 5)
            fill_template_xlsx_row(template_file_path, 21, full_df_balance['Y'].tolist(), 5)

        except Exception as e:
            print(f"An error occurred while filling the PRB template for {cluster_name}: {e}")
            return False

    def generate_prb_for_cluster(self, cluster_name):
        try:
            #create output folder if it doesn't exist
            if not os.path.exists(self.output_subfolder):
                makedir(self.output_subfolder)
            if not os.path.exists(os.path.join(self.output_subfolder, cluster_name)):
                makedir(os.path.join(self.output_subfolder, cluster_name))
            # copy the template file to the output folder and add the cluster name to the file name
            template_file_path = os.path.join(PRB_TEMPLATES_FOLDER, PRB_TEMPLATE)
            shutil.copy(template_file_path, os.path.join(self.output_subfolder, cluster_name, PRB_TEMPLATE))
            # create the new name
            new_file_name = f"{PRB_TEMPLATE.split('.')[0]}{datetime.datetime.now().strftime('%m_%d_%Y')}.xlsx"
            # rename the copied template
            new_file_name = f"{PRB_TEMPLATE.split('.')[0]}{cluster_name}_{datetime.datetime.now().strftime('%m_%d_%Y')}.xlsx"
            new_file_path = os.path.join(self.output_subfolder, cluster_name, new_file_name)
            os.rename(os.path.join(self.output_subfolder, cluster_name, PRB_TEMPLATE), new_file_path)

            self.fill_cluster_dfs(new_file_path, cluster_name)

        except Exception as e:
            print(f"An error occurred while generating the PRB for {cluster_name}: {e}")
            return False

    def generate_prb_files(self):
        was_successful = True
        try:
            self.get_input_files()
            #self.print_input_files() # DEBUG delete later
            print("Generating PRB output...")
            self.generate_prb_for_cluster("ALL")

            # iterate over the cluster names and generate the PRB for each cluster
            cluster = Cluster.get_instance()
            cluster_names = cluster.cluster_names
            for cluster_name in cluster_names:
                self.generate_prb_for_cluster(cluster_name)

            print("PRB output generation completed.")

        except Exception as e:
            print(f"An error occurred while generating the PRB files: {e}")
            was_successful = False
        return was_successful

class Footprint:

    _instance = None
    def __init__(self):
        if Footprint._instance is not None:
            raise Exception("Footprint is a singleton class. Use get_instance() method to access the instance.")
        date = datetime.date.today().strftime("%m_%d_%Y")
        makedir(OUTPUT_FOLDER + "/Huella_" + date)
        self.output_subfolder = f"{OUTPUT_FOLDER}/Huella_{date}"
        self.current_cluster_folder = ""
        self.site_list = []
        self.cluster_name = ""
        self.umts_3g = pd.DataFrame()
        self.filtered3g = pd.DataFrame()
        self.lte_4g = pd.DataFrame()
        self.filtered4g = pd.DataFrame()
        self.nr_5g = pd.DataFrame()
        self.filtered5g = pd.DataFrame()
        self.thor = None
        self.filtered_thor = pd.DataFrame()
        self.ct = None
        self.filteredct = pd.DataFrame()
        self.db = None
        self.filtereddb = pd.DataFrame()
        self.filtereddb_no_dupes = pd.DataFrame()

    @staticmethod
    def get_instance():
        if Footprint._instance is None:
            Footprint._instance = Footprint()
        return Footprint._instance

    def obtain_cluster_footprint(self, cluster_name):
        self.cluster_name = cluster_name
        if UMTS_3G_FILE_PATH != "":
            self.umts_3g = pd.read_csv(UMTS_3G_FILE_PATH, sep=';')
        if LTE_4G_FILE_PATH != "":
            self.lte_4g = pd.read_csv(LTE_4G_FILE_PATH, sep=';')
        if NR_5G_FILE_PATH != "":
            self.nr_5g = pd.read_csv(NR_5G_FILE_PATH, sep=';')
        self.thor = pd.read_csv(THOR_FILE_PATH, sep=';', low_memory=False)
        self.ct = pd.read_csv(CELLTABLE_FILE_PATH, sep=';', low_memory=False)
        self.db = pd.read_csv(DB_FOOTPRINT_FILE_PATH, sep=';')
        self.current_cluster_folder = self.output_subfolder + "/" + cluster_name
        makedir(self.current_cluster_folder)
        iom_instance = IOM.get_instance()
        iom_instance.generate_cluster_sites()
        self.site_list = iom_instance.cells_by_cluster[iom_instance.cells_by_cluster['CLUSTER_NAME'] == cluster_name]['CELL_LIST'].tolist()
        if self.site_list:
            self.site_list = self.site_list[0].split(',')
        self.filter_sites()
        self.generate_output()

    def filter_3g(self):
        filtered_df = pd.DataFrame()
        shorthand = {'VAL': 'V', 'BAL': 'B', 'MUR': 'U', 'AND': 'A', 'EXT': 'E', 'ARA': 'R', 'CYM': 'K', 'CLM': 'X',
                    'MAD': 'M', 'CAN': 'T'}
        print("LISTING 3G:")
        print(self.site_list)
        for target_site in self.site_list:
            for key, value in shorthand.items():
                if target_site.startswith(key):
                    prefix = value
            suffix = ""
            for char in target_site:
                if char.isdigit() and len(suffix) < 4:
                    suffix += char
            pattern = f"{prefix}{suffix}"
            print(f"Prefix: {prefix}, Suffix: {suffix}, Pattern: {pattern}")
            column_name = "Cell Name"
            iteration_df = self.umts_3g[self.umts_3g[column_name].str.contains(pattern, regex=True)]
            filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filtered3g = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTERED3G.csv', sep=';', index=False)

    def filter_4g(self):
        filtered_df = pd.DataFrame()
        print("LISTING 4G:")
        print(self.site_list)
        for target_site in self.site_list:
            prefix = ''.join(char for char in target_site if char.isalpha())
            suffix = ''.join(char for char in target_site if char.isdigit())
            pattern = f"{prefix}X{suffix}"
            column_name = "Cell Name"
            print(f"Prefix: {prefix}, Suffix: {suffix}, Pattern: {pattern}")
            iteration_df = self.lte_4g[self.lte_4g[column_name].str.contains(pattern, regex=True)]
            filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filtered4g = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTERED4G.csv', sep=';', index=False)

    def filter_5g(self):
        filtered_df = pd.DataFrame()
        print("LISTING 5G:")
        print(self.site_list)
        for target_site in self.site_list:
            prefix = ''.join(char for char in target_site if char.isalpha())
            suffix = ''.join(char for char in target_site if char.isdigit())
            pattern = f"{prefix}X{suffix}"
            column_name = "Cell Name"
            print(f"Prefix: {prefix}, Suffix: {suffix}, Pattern: {pattern}")
            iteration_df = self.nr_5g[self.nr_5g[column_name].str.contains(pattern, regex=True)]
            filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filtered5g = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTERED5G.csv', sep=';', index=False)

    def filter_thor(self):
        filtered_df = pd.DataFrame()
        print("LISTING THOR:")
        print(self.site_list)
        for target_site in self.site_list:
            column_name = 'node'
            print(f"Match: {target_site}")
            iteration_df = self.thor[self.thor[column_name].str.contains(target_site)]
            filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filteredthor = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTEREDTHOR.csv', sep=';', index=False)

    def filter_celltable(self):
        filtered_df = pd.DataFrame()
        print("LISTING CELLTABLE:")
        print(self.site_list)
        for target_site in self.site_list:
            column_name = 'SITE'
            print(f"Match: {target_site}")
            iteration_df = self.ct[self.ct[column_name].str.contains(target_site)]
            filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filteredct = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTEREDCT.csv', sep=';', index=False)

    def filter_db(self):
        filtered_df = pd.DataFrame()
        print("LISTING DB:")
        print(self.site_list)
        print(self.cluster_name)
        column_name = 'CLUSTER'
        iteration_df = self.db[self.db[column_name].str.contains(self.cluster_name)]
        filtered_df = pd.concat([filtered_df, iteration_df], ignore_index=True)
        self.filtereddb = filtered_df
        filtered_df.to_csv(self.output_subfolder + '/FILTEREDDB.csv', sep=';', index=False)

    def remove_db_copies(self):
        dupes = self.filtereddb[self.filtereddb.duplicated(subset=['CELLNAMEX'], keep=False)]

        if not dupes.empty:
            print(f"{len(dupes)} dupes found.")
            self.filtereddb['DATE'] = pd.to_datetime(self.filtereddb['DATE'], format='%d/%m/%Y')
            self.filtereddb = self.filtereddb.sort_values(['CELLNAMEX', 'DATE'], ascending=[True, False])
            self.filtereddb = self.filtereddb.drop_duplicates(subset=['CELLNAMEX'], keep='first')
            self.filtereddb['DATE'] = self.filtereddb['DATE'].dt.strftime('%d/%m/%Y')
            self.filtereddb.to_csv(self.output_subfolder + '/FILTEREDDB.csv', sep=';', index=False)
            dupes.to_csv(self.output_subfolder + '/FILTEREDDB_NO_DUPES.csv', sep=';', index=False)

        if dupes.empty:
            print("No dupes found.")
            self.filtereddb_no_dupes = self.filtereddb

    def filter_sites(self):
        if not self.umts_3g.empty:
            self.filter_3g()
        if not self.lte_4g.empty:
            self.filter_4g()
        if not self.nr_5g.empty:
            self.filter_5g()
        self.filter_thor()
        self.filter_celltable()
        self.filter_db()
        self.remove_db_copies()

    def fill_sector_column(self, sheet, sheet_name, data, rows):
        sector_column = None
        cell_name_column = None
        use_last_char = False
        match sheet_name:
            case "3G":
                sector_column = 73
                cell_name_column = 6
                use_last_char = True
            case "4G":
                sector_column = 62
                cell_name_column = 4
            case "5G":
                sector_column = 37
                cell_name_column = 4
            case _:
                print("what.")
                return
        for row_idx in range(rows):
            if use_last_char:
                sheet.cell(row=row_idx + 2, column=sector_column).value = data[row_idx][cell_name_column][-1]
            else:
                sheet.cell(row=row_idx + 2, column=sector_column).value = data[row_idx][cell_name_column][-2]
                pass

    def fill_extra_ct_columns(self, sheet, data, rows):
        band_column = 35
        sector_column = 36
        site_sector_column = 37
        site_column = 9
        cellnamex_column = 11
        for row_idx in range(rows):
            sheet.cell(row=row_idx + 2, column=band_column).value = data[row_idx][cellnamex_column][-3]
            if not data[row_idx][cellnamex_column][-2].isdigit():
                construct = data[row_idx][site_column] + "-" + data[row_idx][cellnamex_column][-1]
                sheet.cell(row=row_idx + 2, column=sector_column).value = data[row_idx][cellnamex_column][-1]
            else:
                construct = data[row_idx][site_column] + "-" + data[row_idx][cellnamex_column][-2]
                sheet.cell(row=row_idx + 2, column=sector_column).value = data[row_idx][cellnamex_column][-2]
            sheet.cell(row=row_idx + 2, column=site_sector_column).value = construct

    def fill_sheet(self, workbook, target_sheet, data, dir, x_offset, data_x_offset=0):
        sheet = workbook[target_sheet]
        num_rows = len(data)
        num_cols = len(data[0]) if num_rows > 0 else 0
        for col_idx in range(num_cols - data_x_offset):
            for row_idx in range(num_rows):
                sheet.cell(row=row_idx + 2, column=col_idx + x_offset, value=data[row_idx][col_idx + data_x_offset])
            print(f"[{target_sheet}]Copying column {col_idx} to the template...")
        if (target_sheet != "CT" or target_sheet != "DB" or target_sheet != "THOR"):
            self.fill_sector_column(sheet, target_sheet, data, num_rows)
        elif (target_sheet == "CT"):
            self.fill_extra_ct_columns(sheet, data, num_rows)
        workbook.save(dir)
        workbook.close()

    def generate_output(self):
        working_file_dir = self.current_cluster_folder + '/Data_Footprint_' + self.cluster_name + '.xlsx'
        shutil.copy(DATA_FOOTPRINT_TEMPLATE_FILE_PATH, working_file_dir)
        workbook = openpyxl.load_workbook(working_file_dir)
        if not self.umts_3g.empty:
            self.fill_sheet(workbook, "3G", self.filtered3g.values.tolist(), working_file_dir, 2)
        if not self.lte_4g.empty:
            self.fill_sheet(workbook, "4G", self.filtered4g.values.tolist(), working_file_dir, 2)
        if not self.nr_5g.empty:
            self.fill_sheet(workbook, "5G", self.filtered5g.values.tolist(), working_file_dir, 2)
        self.fill_sheet(workbook, "Thor", self.filteredthor.values.tolist(), working_file_dir, 1, 1)
        self.fill_sheet(workbook, "CT", self.filteredct.values.tolist(), working_file_dir, 1)
        self.fill_sheet(workbook, "DB", self.filtereddb.values.tolist(), working_file_dir, 1)
        self.fill_sheet(workbook, "DB_filtered", self.filtereddb_no_dupes.values.tolist(), working_file_dir, 1)


# TKINTER ---------------------------------------------------------------------------
class ErrorDialogue(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title = "Error"
        self.geometry("300x60")
        self.create_widgets()

    def create_widgets(self):
        for i in range(3):
            self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(i, weight=1)

        label = tk.Label(self, text="One or more mandatory files (*) are missing.")
        label.grid(row=1, columnspan=3, sticky="nwe", padx=20)
        button = ttk.Button(self, text="OK", command=self.destroy)
        button.grid(row=2, columnspan=3, sticky="e", padx=3, pady=3)

class CellViewer(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title = "Cell Viewer"
        self.geometry("600x475")
        self.create_widgets()

    def create_widgets(self):
        std_button_padding = (3, 1)

        for i in range(9):
            self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(i, weight=1)

        label = ttk.Label(self, text="Cell Viewer | Total Cells")
        label.grid(columnspan=20, row=0)

        title_separator = ttk.Separator(self, orient=tk.HORIZONTAL)
        title_separator.grid(columnspan=20, row=1, sticky='ew')

        button_switch_to_per_cluster = ttk.Button(self, text="Switch to per cluster view", padding= std_button_padding, command=self.on_switch_view)
        button_switch_to_per_cluster.grid(column=0, row=2, padx=20, pady=5, sticky='sw')

        label_ericsson = tk.Label(self, text="Ericsson Cells")
        label_ericsson.grid(column=0, row=3, sticky='sw', padx=20)

        button_copy_ericsson = ttk.Button(self, text="Copy", padding=std_button_padding)
        button_copy_ericsson.grid(column=0, row=3, sticky='se', padx=20)

        display_ericsson = tk.Text(self, height=10, width=80)
        display_ericsson.grid(column=0, row=4, padx=20)
        display_ericsson.insert(tk.END, ERICSSON_LIST)
        display_ericsson.state = tk.DISABLED

        label_huawei = tk.Label(self, text="Huawei Cells")
        label_huawei.grid(column=0, row=5, sticky='sw', padx=20)

        button_copy_huawei = ttk.Button(self, text="Copy", padding=std_button_padding)
        button_copy_huawei.grid(column=0, row=5, sticky='se', padx=20)

        display_huawei = tk.Text(self, height=10, width=80, state=tk.DISABLED)
        display_huawei.grid(column=0, row=6, padx=20)

    def on_switch_view(self):
        pass

class OverwriteDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title = "Huellers 1.0.0"
        self.geometry("350x150")
        self.create_widgets()

    def nuke_output_directory(self):
        # Remove the directory and all its contents.
        try:
            # Call GC first. Under certain circumstances, certain files might still be open and cause issues.
            # This is a workaround for now.
            gc.collect()
            shutil.rmtree(OUTPUT_FOLDER)
            makedir(OUTPUT_FOLDER)
        except:
            print("Directory doesn't exist. Creating...")
            makedir(OUTPUT_FOLDER)

    def create_widgets(self):
        for i in range(3):
            self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(i, weight=1)

        label = ttk.Label(self, text="WARNING: Output directory already contains files!\nDo you wish to overwrite it?\n(Will result in loss of previous data.)")
        label.grid(columnspan=3, row=0, sticky='n', padx=20, pady=20)

        button_yes = ttk.Button(self, text="Yes", command=self.on_yes)
        button_yes.grid(column=0, row=1, sticky='w', padx=20)

        button_no = ttk.Button(self, text="No", command=self.on_no)
        button_no.grid(column=2, row=1, sticky='e', padx=20)

    def on_yes(self):
        print("Overwrite confirmed.")
        self.nuke_output_directory()
        self.determine_procedure()
        self.master.destroy()
        self.destroy()

    def on_no(self):
        print("Overwrite denied.")
        self.destroy()

class FootProgressReporter(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title = "Huellers 1.0.0"
        self.geometry("450x100")
        self.total_clusters = 0
        self.current_cluster = 0
        self.processing_text = None
        self.progress_widget = None
        self.create_widgets()

    def process(self):
        fp = Footprint.get_instance()
        total_clusters = USER_CLUSTERS.split('\n')
        curr = 2
        for cluster in total_clusters:
            self.processing_text.set(f"Processing cluster {curr - 1} of {len(total_clusters) - 1}: {cluster}")
            fp.obtain_cluster_footprint(cluster)
            curr += 1
            sleep(1)

    def create_widgets(self):
        self.processing_text = tk.StringVar()
        self.processing_text.set(f"Processing cluster {self.current_cluster} of {self.total_clusters}...")

        for i in range(3):
            self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(i, weight=1)
        self.progress_widget = tk.Label(self, textvariable=self.processing_text)
        self.progress_widget.grid(columnspan=3, row=1, sticky='nw', padx=20)
        label_disclaimer = ttk.Label(self, text="This might take several minutes and the app might become unresponsive.\nPlease be patient.")
        label_disclaimer.grid(columnspan=3, row=2, sticky='sw', padx=20, pady=20)

class FileRequester(tk.Toplevel):

    def __init__(self, parent):
        super().__init__(parent)
        self.title = "Seleccionar Archivos"
        self.mode = "default"
        match EXPORT_MODE:
            case "IOM tmp":
                self.geometry("400x200")
            case "PRBs + Footprint":
                self.geometry("400x400")
            case _:
                self.geometry("400x200")
        self.create_widgets()

    def open_csv_file(self, filename, label):
        file = filedialog.askopenfilename(title="Select a csv file, please.", filetypes=(("CSV files", "*.csv"), ("All files", "*.*")))
        if not file:
            return
        match filename:
            case "Tracking CI":
                global QUALITY_BOOST_TRACKING_FILE_PATH
                QUALITY_BOOST_TRACKING_FILE_PATH = file
            case "Celltable":
                global CELLTABLE_FILE_PATH
                CELLTABLE_FILE_PATH = file
            case "Clusters":
                pass
            case "Thor":
                global THOR_FILE_PATH
                THOR_FILE_PATH = file
            case "DB Footprint":
                global DB_FOOTPRINT_FILE_PATH
                DB_FOOTPRINT_FILE_PATH = file
            case "3G":
                global UMTS_3G_FILE_PATH
                UMTS_3G_FILE_PATH = file
            case "4G":
                global LTE_4G_FILE_PATH
                LTE_4G_FILE_PATH = file
            case "5G":
                global NR_5G_FILE_PATH
                NR_5G_FILE_PATH = file
            case _:
                print("Something has gone terribly wrong.")
        label.set(filename + (" (Status: submitted)"))
        print(label)

    def determine_procedure(self, ericsson):
        # This function determines which procedure to run based on the selected EXPORT_MODE.
        match EXPORT_MODE:
            case "IOM tmp":
                iom = IOM.get_instance()
                iom.generate_tmp_iom_files()
            case "PRBs + Footprint":
                prog = FootProgressReporter(self)
                prog.grab_set()
                process_thread = threading.Thread(target=prog.process)
                process_thread.daemon = True
                process_thread.start()
                if ericsson:
                    eric = AdapterEricsson.get_instance()
                    eric.generate_new_input()
                prb = PRB.get_instance()
                prb.generate_prb_files()


    def on_process_files(self):
        print(OUTPUT_FOLDER)
        """if QUALITY_BOOST_TRACKING_FILE_PATH == "":
            error_dialogue = ErrorDialogue(self)
            error_dialogue.grab_set()
            return
        if CELLTABLE_FILE_PATH == "":
            error_dialogue = ErrorDialogue(self)
            error_dialogue.grab_set()
            return
        if EXPORT_MODE == "PRBs + Footprint":
            g = False
            if THOR_FILE_PATH == "":
                error_dialogue = ErrorDialogue(self)
                error_dialogue.grab_set()
                return
            if DB_FOOTPRINT_FILE_PATH == "":
                error_dialogue = ErrorDialogue(self)
                error_dialogue.grab_set()
                return
            if UMTS_3G_FILE_PATH == "":
                g = True
            if LTE_4G_FILE_PATH == "":
                g = True
            if NR_5G_FILE_PATH == "":
                g = True
            if not g:
                error_dialogue = ErrorDialogue(self)
                error_dialogue.grab_set()
                return"""
        # If the output folder is empty, proceed normally.
        if len(os.listdir(OUTPUT_FOLDER)) == 0:
            self.determine_procedure()
        # If it contains anything, warn the user.
        else:
            overwrite_dialog = OverwriteDialog(self)
            overwrite_dialog.grab_set()

    # Creates the widgets for the PRB + Footprint mode
    def create_prb_footprint_widgets(self, row_start):
        std_button_padding = (3, 1)
        text_thor_cell_scoring_label = tk.StringVar()
        text_thor_cell_scoring_label.set("Thor Celltable* (Status: missing)")
        text_db_footprint_label = tk.StringVar()
        text_db_footprint_label.set("DB Footprint* (Status: missing)")
        text_3g_label = tk.StringVar()
        text_3g_label.set("3G [UMTS] (Status: missing)")
        text_4g_label = tk.StringVar()
        text_4g_label.set("4G [LTE] (Status: missing)")
        text_5g_label = tk.StringVar()
        text_5g_label.set("5G [NR] (Status: missing)")

        label_thor_cell_scoring = tk.Label(self, textvariable=text_thor_cell_scoring_label)
        label_thor_cell_scoring.grid(column=0, row=row_start + 1, sticky='wn', padx=20)

        button_browse_thor_cell_scoring = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("Thor", text_thor_cell_scoring_label))
        button_browse_thor_cell_scoring.grid(column=4, row=row_start + 1, sticky='ne', padx=20)

        label_db_footprint = tk.Label(self, textvariable=text_db_footprint_label)
        label_db_footprint.grid(column=0, row=row_start + 2, sticky='wn', padx=20)

        button_browse_db_footprint = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("DB Footprint", text_db_footprint_label))
        button_browse_db_footprint.grid(column=4, row=row_start + 2, sticky='ne', padx=20)

        label_3g = tk.Label(self, textvariable=text_3g_label)
        label_3g.grid(column=0, row=row_start + 3, sticky='wn', padx=20)

        button_browse_3g = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("3G", text_3g_label))
        button_browse_3g.grid(column=4, row=row_start + 3, sticky='ne', padx=20)

        label_4g = tk.Label(self, textvariable=text_4g_label)
        label_4g.grid(column=0, row=row_start + 4, sticky='wn', padx=20)

        button_browse_4g = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("4G", text_4g_label))
        button_browse_4g.grid(column=4, row=row_start + 4, sticky='ne', padx=20)

        label_5g = tk.Label(self, textvariable=text_5g_label)
        label_5g.grid(column=0, row=row_start + 5, sticky='wn', padx=20)

        button_browse_5g = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("5G", text_5g_label))
        button_browse_5g.grid(column=4, row=row_start + 5, sticky='ne', padx=20)

    def create_widgets(self):
        std_button_padding = (3, 1)
        text_tracking_ci_label = tk.StringVar()
        text_tracking_ci_label.set("Tracking CI* (Status: missing)")
        text_tracking_celltable_label = tk.StringVar()
        text_tracking_celltable_label.set("Celltable* (Status: missing)")

        if EXPORT_MODE == "default" or EXPORT_MODE == "IOM tmp":
            for i in range(5):
                self.grid_columnconfigure(i, weight=1)
                self.grid_rowconfigure(i, weight=1)
        elif EXPORT_MODE == "PRBs + Footprint":
            for i in range(10):
                self.grid_rowconfigure(i, weight=1)
                if i < 5:
                    self.grid_columnconfigure(i, weight=1)

        label = ttk.Label(self, text="Select files")
        label.grid(columnspan=5, row=0, sticky='n', pady=5)

        title_separator = ttk.Separator(self, orient=tk.HORIZONTAL)
        title_separator.grid(columnspan=5, row=0, sticky='ew')

        label_tracking_ci = tk.Label(self, textvariable=text_tracking_ci_label)
        label_tracking_ci.grid(column=0, row=1, sticky='wn', padx=20)

        button_browse_tracking_ci = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("Tracking CI", text_tracking_ci_label))
        button_browse_tracking_ci.grid(column=4, row=1, sticky='ne', padx=20)

        label_celltable = tk.Label(self, textvariable=text_tracking_celltable_label)
        label_celltable.grid(column=0, row=2, sticky='wn', padx=20)

        button_browse_celltable = ttk.Button(self, text="Open", padding=std_button_padding, command=lambda: self.open_csv_file("Celltable", text_tracking_celltable_label))
        button_browse_celltable.grid(column=4, row=2, sticky='ne', padx=20)

        if EXPORT_MODE == "PRBs + Footprint":
            self.create_prb_footprint_widgets(2)

        button_ok = ttk.Button(self, text="Process Files", padding=std_button_padding, command=lambda: self.determine_procedure(False))
        if (EXPORT_MODE == "default" or EXPORT_MODE == "IOM tmp"):
            button_ok.grid(columnspan=5, row=3, sticky='we', padx=20)
        elif EXPORT_MODE == "PRBs + Footprint":
            button_ok_ericsson = ttk.Button(self, text="Process Files (Ericcson)", padding=std_button_padding, command=lambda: self.determine_procedure(True))
            button_ok_ericsson.grid(columnspan=5, row=8, sticky='we', padx=20)
            button_ok.grid(columnspan=5, row=9, sticky='we', padx=20)

# Utils ------------------------------------------------------------------
def fill_template_xlsx(file_path, position, values, sheet_name=None):
    # fills the template in the column position with the values in the list values
    try:
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        for index, value in enumerate(values):
            sheet.cell(row=index+2, column=position, value=value)
        workbook.save(file_path)
        workbook.close()
    except Exception as e:
        print(f"An error occurred: {e}")

def fill_template_xlsx_row(file_path, position, values, starting_row):
    # fills the template in the column position with the values in the list values
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for index, value in enumerate(values):
            sheet.cell(row=index+starting_row, column=position, value=value)
        workbook.save(file_path)
        workbook.close()
    except Exception as e:
        print(f"An error occurred: {e}")

def clear_folder(path):
    # removes every file and subfolder from the path
    try:
        for root, dirs, files in os.walk(path, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
    except Exception as e:
        print(f"An error occurred while clearing the folder: {e}")

def remove_out_of_date_csv_rows(file_path, date_column):
    # removes every row that is not in the valid date range
    start = DATE_START
    end = DATE_END
    date_format = '%Y-%m-%d %H:%M'
    try:
        df = pd.read_csv(file_path, sep=';')
        df[date_column] = pd.to_datetime(df[date_column], format=date_format)
        mask = (df[date_column] >= start) & (df[date_column] <= end)
        df = df.loc[mask]
        df.to_csv(file_path, sep=';', index=False)
    except Exception as e:
        print(f"An error occurred: {e}")

def makedir(path):
    # creates the directory if it doesn't exist
    try:
        if not os.path.exists(path):
            os.makedirs(path)
    except Exception as e:
        print(f"An error occurred while creating the directory: {e}")

# Tests -------------------------------------------------------------------------

def print_csv_head(file_path):
    try:
        df = pd.read_csv(file_path, sep=';')
        print(df.head())
    except Exception as e:
        print(f"An error occurred: {e}")

def test_cluster():
    try:
        cluster = Cluster.get_instance()
        #print the list of cluster names
        print("Cluster names:", cluster.cluster_names)
        # test the singleton property
        cluster2 = Cluster.get_instance()
        if cluster is cluster2:
            print("Singleton test passed.")
        else:
            print("Singleton test failed.")

    except Exception as e:
        print(f"Cluster instance test failed: {e}")

def test_csv_match(file_path, value):
    # test in which columns the value is found in the csv file
    try:
        df = pd.read_csv(file_path, sep=';')
        # check if the value is in the dataframe
        test_dataframe_match(df, value)
    except Exception as e:
        print(f"An error occurred: {e}")

# use this to check where to filter in the dataframe
def test_dataframe_match(dataframe, value):
    # test in which columns the value is found in the dataframe
    try:
        match_list = []
        # check if the value is in the dataframe
        if value in dataframe.values:
            print(f"Value {value} found in dataframe.")
            # get the column name where the value is found
            for column in dataframe.columns:
                if value in dataframe[column].values:
                    match_list.append(column)
                    print(f"Value {value} found in column {column}.")
        else:
            print(f"Value {value} not found in dataframe.")
    except Exception as e:
        print(f"An error occurred: {e}")


def normalize_value(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    # Try to convert "3.14" to "3,14"
    # the decimal length is indeterminate, so we need to check if the string is in the format of "3,14" or "-3,14"
    if re.match(r'^[+-]?\d+(\.\d+)?$', val_str):
        val_str = val_str.replace('.', ',')
    # Return as float if possible, else string
    try:
        return float(val_str)
    except ValueError:
        return val_str

def test_find_matching_columns(xlsx_path, dataframes):
    # Load Excel template (first sheet)
    template_df = pd.read_excel(xlsx_path)
    unmatched_columns_numeric = []

    print("Matching columns in the Excel template:\n")
    # print the head of the template xlsx file
    print(template_df.head())

    for template_col in template_df.columns:
        template_values = set(template_df[template_col].dropna().apply(normalize_value))
    #print the normalized template values, first 5 values
        # check if the column is empty
        if template_values == set():
            print(f"Template column '{template_col}' is empty.")
            continue
        for i, df in enumerate(dataframes):
            matched = False
            for df_col in df.columns:
                df_values = set(df[df_col].dropna().astype(str))
                if template_values & df_values:
                    print(f"- Template column '{template_col}' matches values from DataFrame #{i + 1}, column '{df_col}'")
                    matched = True
            if not matched:
                unmatched_columns_numeric.append(template_col)

    print("\nUnmatched columns in the template:")
    for col in unmatched_columns_numeric:
        print(f"- {col}")


def test_dataframe_columns(dataframe, columns):
    # test if the columns are in the dataframe
    try:
        for column in columns:
            if column not in dataframe.columns:
                print(f"Column {column} not found in dataframe.")
            else:
                print(f"Column {column} found in dataframe.")
    except Exception as e:
        print(f"An error occurred: {e}")

def test_if_columns_in_dataframes(dataframes, columns):
    # prints the matching dataframe for each column in the columns list
    # it is not case sensitive
    try:
        for column in columns:
            for i, df in enumerate(dataframes):
                if column.lower() in df.columns.str.lower():
                    print(f"Column {column} found in dataframe {i + 1}.")
                else:
                    print(f"Column {column} not found in dataframe {i + 1}.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Add more tests here
def master_test():
    print("RUNNING MASTER TEST")
    test_cluster()
    # Test CSV processing
    iom = IOM.get_instance()
    iom.generate_cluster_sites()
    iom.generate_tmp_iom_files()
    print("MASTER TEST FINISHED")

def test_input_files():
    try:
        if not os.path.exists(QUALITY_BOOST_TRACKING_FILE_PATH):
            print("Quality Boost Tracking file not found.")
            return False
        if not os.path.exists(CELLTABLE_FILE_PATH):
            print("Celltable file not found.")
            return False
        if not os.path.exists(THOR_FILE_PATH):
            print("Thor file not found.")
            return False
        if not os.path.exists(DB_FOOTPRINT_FILE_PATH):
            print("DB Footprint file not found.")
            return False
        if not os.path.exists(UMTS_3G_FILE_PATH):
            print("3G file not found.")
            return False
        if not os.path.exists(LTE_4G_FILE_PATH):
            print("4G file not found.")
            return False
        if not os.path.exists(NR_5G_FILE_PATH):
            print("5G file not found.")
            return False
        if not os.path.exists(PRB_TEMPLATES_FOLDER):
            print("PRB Templates folder not found.")
            return False
        if not os.path.exists(DATA_FOOTPRINT_TEMPLATE_FILE_PATH):
            print("Data Footprint Template file not found.")
            return False
        if not os.path.exists(CLUSTERS_FILE_PATH):
            print("Clusters file not found.")
            return False
        return True
    except Exception as e:
        print(f"An error occurred: {e}")
        return False

# tkinter UI -------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Huellers 1.0.0")
        self.geometry("400x300")

        # TTK VARIABLE DECLARATIONS
        self.checked_iom_tmp = tk.BooleanVar()
        self.checked_iom_tmp.set(False)
        # TTK VARIABLE DECLARATIONS END
        self.create_widgets()

    def open_file_requester(self):
        file_requester = FileRequester(self)
        file_requester.grab_set()

    def create_widgets(self):
        std_button_padding = (3,1)
        export_options = ["IOM tmp", "PRBs + Footprint"]
        # Configure UI grid.
        for i in range(9):
            self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(i, weight=1)

        program_title_label = ttk.Label(self, text="Huellers Tool Redux")
        program_title_label.grid(columnspan=9, row=0)

        title_separator = ttk.Separator(self, orient='horizontal')
        title_separator.grid(columnspan=9, row=1, sticky='ew')

        label_clusters = ttk.Label(self, text="Selected Clusters (one per line, scrolls):")
        label_clusters.grid(columnspan=9, row=2, sticky='w', padx=20)

        paste_button = ttk.Button(self, text="Paste", padding=0, command= lambda: self.on_paste_clusters(cluster_entry))
        paste_button.grid(column=2, row=2, sticky='e', padx=20)

        cluster_entry = tk.Text(self, height=5, width=50)
        cluster_entry.grid(columnspan=9, row=3, sticky='nw', padx=20)

        label_export_mode = ttk.Label(self, text="Export Mode:")
        label_export_mode.grid(column=0, row=4, sticky='nw', padx=20)

        dropdown_export_mode = ttk.Combobox(self, values=export_options, state='readonly')
        dropdown_export_mode.grid(column=1, row=4, sticky='nw', padx=3)
        dropdown_export_mode.current(0)

        options_separator = ttk.Separator(self, orient='horizontal')
        options_separator.grid(columnspan=9, row=5, sticky='ew')

        label_operations_section = ttk.Label(self, text="Operations")
        label_operations_section.grid(column=0, row=6, sticky='nw', padx=20)

        export_clusters_button = ttk.Button(self, text="Export clusters", width=15, padding=std_button_padding, command= lambda: self.on_export_clusters(dropdown_export_mode, cluster_entry))
        export_clusters_button.grid(column=0, row=6, sticky='sw', padx=20, pady=(15, 0))

        view_cells_button = ttk.Button(self, text="View cell lists", width=15, padding=std_button_padding, command= self.on_view_cells)
        view_cells_button.grid(column=1, row=6, sticky='sw', padx=0, pady=(15, 0))

    def on_view_cells(self):
        cell_viewer = CellViewer(self)
        cell_viewer.grab_set()

    def on_export_clusters(self, mode, cluster_entry):
        global EXPORT_MODE
        EXPORT_MODE = mode.get()
        global USER_CLUSTERS
        USER_CLUSTERS = cluster_entry.get('1.0', tk.END)
        cluster = Cluster.get_instance()
        cluster.get_cluster_names()
        print(EXPORT_MODE)
        print(cluster_entry.get('1.0', tk.END))
        self.open_file_requester()

    def on_paste_clusters(self, cluster_entry):
        cluster_entry.delete('1.0', tk.END)
        cluster_entry.insert('1.0', self.clipboard_get())

    def load_csv(self):
        file_path = filedialog.askopenfilename(
            title="Select a CSV File",
            filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*"))
        )
        if file_path:
            print_csv_head(file_path)


# Main and Initialization -----------------------------------------------------

def init():
    is_valid = True
    print("Initializing...")
    #test_input_files()
    #if not test_input_files():
    #    exit(1)
    #cluster = Cluster.get_instance()
    #prb = PRB.get_instance()
    #if not prb.generate_prb_files():
    #    is_valid = False
    #if not is_valid:
    #    print("Initialization failed.")
    #    exit(1)
    #footprint = Footprint.get_instance()
    print("Initialization complete.")


if __name__ == "__main__":
    clear_folder(OUTPUT_FOLDER) # remove
    init()
    #master_test() # this must be commented or removed for release
    app = App()
    app.mainloop()
